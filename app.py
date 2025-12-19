from flask import Flask, render_template, request, redirect, session, jsonify, send_file
from flask_session import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import datetime, date, timedelta
import os
import requests
from zoneinfo import ZoneInfo

app = Flask(__name__)

# =======================================
# CONFIGURACIÓN DE SESIÓN (MEJORADA)
# =======================================
app.secret_key = os.environ.get("SECRET_KEY", "super_secret_key_123")

app.config.update(
    SESSION_TYPE="filesystem",              # Guardar sesión en el servidor
    SESSION_FILE_DIR="./.flask_session",    # Carpeta de sesiones
    SESSION_PERMANENT=True,
    PERMANENT_SESSION_LIFETIME=timedelta(days=1),  # 1 día
    SESSION_USE_SIGNER=True,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax"
)

Session(app)

# =======================================
# ARCHIVO DINÁMICO POR USUARIO + RUTA
# =======================================
def archivo_excel():
    usuario = session.get("usuario", "user").replace(" ", "_")
    ruta = session.get("ruta", "ruta").replace(" ", "_")
    return f"registro_{usuario}_{ruta}.xlsx"

# =======================================
# CATEGORÍAS BASE
# =======================================
categorias_base = [
    "Autos", "Camionetas", "Micro Bus", "Mini Bus", "Bus", "Omnibus",
    "Camiones de 1 Eje", "Camiones de 2 Ejes", "Camiones 3 Ejes o más",
    "Motos", "Jeeps", "Bicicletas", "Peatones", "V.A", "V.C",
    "Tracción Animal", "Rickshaw"
]

# =======================================
# ESTILOS EXCEL
# =======================================
def estilizar(ws):
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border

# =======================================
# CREAR EXCEL INICIAL
# =======================================
def inicializar_excel():
    archivo = archivo_excel()

    if not os.path.exists(archivo):
        wb = openpyxl.Workbook()

        c = wb.active
        c.title = "Conteo"
        c.append(["Categoría", "Conteo", "Fecha", "Ruta"])

        for cat in categorias_base:
            c.append([cat, 0, "", ""])

        h = wb.create_sheet("Historial")
        h.append(["Fecha", "Hora", "Ruta", "Usuario", "Categoría", "Cantidad", "N° Vehículos"])

        estilizar(c)
        estilizar(h)
        wb.save(archivo)

# =======================================
# LOGIN
# =======================================
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form["usuario"].strip()
        ruta = request.form["ruta"].strip()

        if not usuario or not ruta:
            return render_template("login.html", error="Debes completar Usuario y Ruta")

        session.permanent = True  # 🔒 sesión persistente
        session["usuario"] = usuario
        session["ruta"] = ruta
        session["conteos"] = {c: 0 for c in categorias_base}
        session["nuevas"] = {}
        session["vehiculos_hoy"] = 0

        inicializar_excel()
        return redirect("/contador")

    return render_template("login.html")

# =======================================
# CONTADOR
# =======================================
@app.route("/contador")
def contador():
    if "usuario" not in session:
        return redirect("/")

    return render_template(
        "contador.html",
        categorias_base=categorias_base,
        conteos=session["conteos"],
        nuevas=session["nuevas"]
    )

# =======================================
# MODIFICAR CONTEO
# =======================================
@app.route("/modificar", methods=["POST"])
def modificar():
    cat = request.form["categoria"]
    accion = request.form["accion"]

    if cat in session["conteos"]:
        if accion == "sumar":
            session["conteos"][cat] += 1
            session["vehiculos_hoy"] += 1
        elif accion == "restar" and session["conteos"][cat] > 0:
            session["conteos"][cat] -= 1
            session["vehiculos_hoy"] -= 1
    else:
        if accion == "sumar":
            session["nuevas"][cat] += 1
            session["vehiculos_hoy"] += 1
        elif accion == "restar" and session["nuevas"][cat] > 0:
            session["nuevas"][cat] -= 1
            session["vehiculos_hoy"] -= 1

    session.modified = True
    return jsonify(ok=True)

# =======================================
# NUEVA CATEGORÍA
# =======================================
@app.route("/nueva_categoria", methods=["POST"])
def nueva_categoria():
    nombre = request.form["nombre"].strip()

    if not nombre:
        return jsonify(error="Debe escribir un nombre")

    if nombre in session["conteos"] or nombre in session["nuevas"]:
        return jsonify(error="La categoría ya existe")

    session["nuevas"][nombre] = 0
    session.modified = True
    return jsonify(ok=True)

# =======================================
# GUARDAR EXCEL
# =======================================
@app.route("/guardar", methods=["POST"])
def guardar():
    archivo = archivo_excel()
    wb = openpyxl.load_workbook(archivo)
    conteo = wb["Conteo"]
    hist = wb["Historial"]

    zona_chile = ZoneInfo("America/Santiago")
    ahora = datetime.now(zona_chile)

    fecha = ahora.strftime("%d-%m-%Y")
    hora = ahora.strftime("%H:%M:%S")

    def actualizar(cat, cant):
        for row in conteo.iter_rows(min_row=2):
            if row[0].value == cat:
                row[1].value += cant
                row[2].value = fecha
                row[3].value = session["ruta"]
                break
        else:
            conteo.append([cat, cant, fecha, session["ruta"]])

        hist.append([
            fecha, hora, session["ruta"], session["usuario"],
            cat, cant, session["vehiculos_hoy"]
        ])

    for c, n in session["conteos"].items():
        if n > 0:
            actualizar(c, n)

    for c, n in session["nuevas"].items():
        if n > 0:
            actualizar(c, n)

    # eliminar total anterior
    for i, row in enumerate(conteo.iter_rows(min_row=2), start=2):
        if row[0].value == "N° Vehículos":
            conteo.delete_rows(i)
            break

    total = sum(
        row[1].value for row in conteo.iter_rows(min_row=2)
        if row[0].value != "N° Vehículos"
    )

    conteo.append(["N° Vehículos", total, fecha, session["ruta"]])

    wb.save(archivo)

    session["conteos"] = {c: 0 for c in categorias_base}
    session["nuevas"] = {c: 0 for c in session["nuevas"]}
    session.modified = True

    return jsonify(ok=True, mensaje="Datos guardados correctamente.")

# =======================================
# DESCARGAR EXCEL
# =======================================
@app.route("/abrir_excel")
def abrir_excel():
    if "usuario" not in session:
        return redirect("/")

    archivo = archivo_excel()
    return send_file(archivo, as_attachment=True)

# =======================================
# TELEGRAM
# =======================================
def enviar_excel_por_telegram(archivo):
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")

    if not token or not chat_id:
        raise Exception("Telegram no configurado")

    url = f"https://api.telegram.org/bot{token}/sendDocument"

    with open(archivo, "rb") as f:
        response = requests.post(
            url,
            data={
                "chat_id": chat_id,
                "caption": "📊 Registro de conteo de vehículos"
            },
            files={
                "document": f
            }
        )

    if response.status_code != 200:
        raise Exception(response.text)

# =======================================
# CERRAR SESIÓN
# =======================================
@app.route("/cerrar", methods=["POST"])
def cerrar():
    archivo = archivo_excel()

    try:
        enviar_excel_por_telegram(archivo)
        session.clear()
        return jsonify(
            ok=True,
            mensaje="📤 Excel enviado correctamente por Telegram."
        )

    except Exception as e:
        return jsonify(
            ok=False,
            mensaje=f"❌ Error enviando por Telegram: {e}"
        )

# =======================================
# RUN
# =======================================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))

